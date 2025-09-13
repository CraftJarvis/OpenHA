import os 
import av
import json
import math
import numpy as np 
import pickle
import rich
from tqdm import tqdm
import os
import shutil
import pathlib
import hashlib
import uuid
from collections import defaultdict
from typing import Union,Optional,Any,List,Dict
from datetime import datetime
import subprocess

def save_render_videos(rollout_path, render_frames, fps=20):
    if len(render_frames) == 0:
        return 
    if rollout_path.endswith('.mp4'):
        output_path = rollout_path
    else:
        output_path = os.path.join(rollout_path, 'render.mp4')
    with av.open(output_path, mode="w", format='mp4') as container:
        stream = container.add_stream("h264", rate=fps)
        stream.width = render_frames[0].shape[1]
        stream.height = render_frames[0].shape[0]
        for idx,frame in enumerate(render_frames):
            frame = av.VideoFrame.from_ndarray(frame, format="rgb24")
            for packet in stream.encode(frame):
                container.mux(packet)
        for packet in stream.encode():
            container.mux(packet)

def clip_action_info(instance_folder, assert_video_action_num_same:bool=False, clip_raw_action:bool=False, clip_info:bool=True):
    """
    由于openha在保存时不会保存第一帧，因此对齐错误。为此，重新对齐
    具体方法是：
    1. 剪掉action第一帧
    
    """
            
    # 首先，获取视频长度
    if not assert_video_action_num_same:
        # 获取action长度
        action_num = 0
        with open(os.path.join(instance_folder, 'action.jsonl'), 'r') as f:
            for _ in f:
                action_num += 1
        
        
        frame_num = 0
        if os.path.exists(os.path.join(instance_folder, 'episode.jsonl')):
            with open(os.path.join(instance_folder, 'episode.jsonl'), 'r') as f:
                for _ in f:
                    frame_num += 1
        elif os.path.exists(os.path.join(instance_folder, 'episode_1.mp4')):
            # load the mp4 
            # import ipdb; ipdb.set_trace()
            with av.open(os.path.join(instance_folder, 'episode_1.mp4'), mode="r", format='mp4') as container:
                for packet in container.demux():
                    for _ in packet.decode():
                        frame_num += 1
        else:
            raise ValueError(f"No episode.jsonl or episode_1.mp4 found in {instance_folder}")
        assert action_num == frame_num
    
    subprocess.run(f"sed -i '1d' {os.path.join(instance_folder, 'action.jsonl')}", shell=True, check=True)
    if clip_info:
        subprocess.run(f"sed -i '1d' {os.path.join(instance_folder, 'info.jsonl')}", shell=True, check=True)
    if clip_raw_action:
        subprocess.run(f"sed -i '1d' {os.path.join(instance_folder, 'raw_action.jsonl')}", shell=True, check=True)
    

def extract_frames_from_mp4(mp4_path):
    frames = []
    with av.open(mp4_path, mode="r", format='mp4') as container:
        for packet in container.demux():
            for frame in packet.decode():
                frames.append(frame.to_ndarray(format="rgb24"))
    return frames


def generate_uuid():
    return str(uuid.uuid4())

def generate_timestamp():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
########################################################################
def load_json_file(file_path: Union[str, pathlib.Path], data_type="dict"):
    """
    Load a JSON file from the given path.

    Args:
        file_path (Union[str, pathlib.Path]): Path to the JSON file.
        data_type (str): Expected data type of the JSON content ("dict" or "list").

    Returns:
        dict or list: Loaded JSON content. Returns an empty dictionary or list if the file does not exist.
    """
    if isinstance(file_path, pathlib.Path):
        file_path = str(file_path)  # Convert pathlib.Path to string

    # Initialize an empty object based on the specified data type
    if data_type == "dict":
        json_file = dict()
    elif data_type == "list":
        json_file = list()
    else:
        raise ValueError("Invalid data type. Expected 'dict' or 'list'.")

    # Check if the file exists
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r', encoding="utf-8") as f:
                json_file = json.load(f)  # Load JSON content
        except IOError as e:
            rich.print(f"[red]Failed to open file {file_path}: {e}[/red]")
        except json.JSONDecodeError as e:
            rich.print(f"[red]Error parsing JSON file {file_path}: {e}[/red]")
    else:
        rich.print(f"[yellow]File {file_path} does not exist. Returning an empty file...[/yellow]")

    return json_file

def dump_json_file(json_file, file_path: Union[str, pathlib.Path], indent=4, if_print=True, if_backup=True, if_backup_delete=False):
    """
    Save data to a JSON file with optional backup and logging.

    Args:
        json_file (dict or list): The JSON data to save.
        file_path (Union[str, pathlib.Path]): Path to save the JSON file.
        indent (int): Indentation level for formatting the JSON file (default is 4).
        if_print (bool): Whether to print status messages (default is True).
        if_backup (bool): Whether to create a backup before writing (default is True).
        if_backup_delete (bool): Whether to delete the backup after writing (default is False).
    """
    if isinstance(file_path, pathlib.Path):
        file_path = str(file_path)  # Convert pathlib.Path to string

    backup_path = file_path + ".bak"  # Define the backup file path

    # Create a backup if the file exists and backup is enabled
    if os.path.exists(file_path) and if_backup:
        shutil.copy(file_path, backup_path)

    before_exist = os.path.exists(file_path)  # Check if the file existed before writing

    try:
        # Write JSON data to file
        with open(file_path, 'w', encoding="utf-8") as f:
            json.dump(json_file, f, indent=indent, ensure_ascii=False)

        # Print status messages
        if before_exist and if_print:
            rich.print(f"[yellow]Updated {file_path}[/yellow]")
        elif if_print:
            rich.print(f"[green]Created {file_path}[/green]")

    except IOError as e:
        # Restore from backup if writing fails
        if os.path.exists(backup_path) and if_backup:
            shutil.copy(backup_path, file_path)
            if if_print:
                rich.print(f"[red]Failed to write {file_path}. Restored from backup: {e}[/red]")
        else:
            if if_print:
                rich.print(f"[red]Failed to write {file_path} and no backup available: {e}[/red]")

    finally:
        # Cleanup: Delete the backup file if required
        if if_backup:
            if os.path.exists(backup_path) and if_backup_delete:
                os.remove(backup_path)
            elif not os.path.exists(backup_path) and not if_backup_delete:  # If the file was initially empty, create a backup
                shutil.copy(file_path, backup_path)

def dump_jsonl(jsonl_file:list,file_path:Union[str , pathlib.Path],if_print=True):
    if isinstance(file_path,pathlib.Path):
        file_path = str(file_path)
    before_exist = os.path.exists(file_path)
    try:
        with open(file_path, 'w',encoding="utf-8") as f:
            for entry in jsonl_file:
                json_str = json.dumps(entry,ensure_ascii=False)
                f.write(json_str + '\n') 
        if before_exist and if_print:
            rich.print(f"[yellow]更新{file_path}[/yellow]")
        elif if_print:
            rich.print(f"[green]创建{file_path}[/green]")
    except IOError as e:
        print(f"[red]文件{file_path}写入失败，{e}[/red]") 

def split_dump_jsonl(jsonl_file:list,file_path:Union[str , pathlib.Path],split_num = 1, if_print=True):
    # 确保 file_path 是字符串类型
    if isinstance(file_path, pathlib.Path):
        file_path = str(file_path)
    
    # 检查原文件是否存在
    before_exist = os.path.exists(file_path)
    
    # 计算每份数据的大小
    chunk_size = len(jsonl_file) // split_num
    remainder = len(jsonl_file) % split_num  # 计算余数，用来均衡每份的大小

    # 将数据切分成 5 份
    chunks = []
    start_idx = 0
    for i in range(split_num):
        end_idx = start_idx + chunk_size + (1 if i < remainder else 0)  # 如果有余数，前几个切片多分一个
        chunks.append(jsonl_file[start_idx:end_idx])
        start_idx = end_idx
    
    # 写入每份数据到文件
    for i, chunk in enumerate(chunks):
        try:
            # 构造文件路径
            chunk_file_path = file_path[:-6] + f"{i}.jsonl"
            with open(chunk_file_path, 'w', encoding="utf-8") as f:
                for entry in chunk:
                    json_str = json.dumps(entry, ensure_ascii=False)
                    f.write(json_str + '\n')

            # 打印文件创建或更新信息
            if before_exist and if_print:
                rich.print(f"[yellow]更新{chunk_file_path}[/yellow]")
            elif if_print:
                rich.print(f"[green]创建{chunk_file_path}[/green]")
        except IOError as e:
            print(f"[red]文件{chunk_file_path}写入失败，{e}[/red]")

def load_jsonl(file_path:Union[str , pathlib.Path]):
    if isinstance(file_path,pathlib.Path):
        file_path = str(file_path)
    jsonl_file = []
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r') as f:
                for line in f:
                    jsonl_file.append(json.loads(line))
        except IOError as e:
            rich.print(f"[red]无法打开文件：{e}")
        except json.JSONDecodeError as e:
            rich.print(f"[red]解析 JSON 文件时出错：{e}")
    else:
        rich.print(f"[yellow]{file_path}文件不存在，正在传入空文件...[/yellow]")
    return jsonl_file 
                
class JsonlProcessor:
    def __init__(self, file_path:Union[str , pathlib.Path],
                 if_backup = True,
                 if_print=True
                 ):
        
        self.file_path = file_path if not isinstance(file_path,pathlib.Path) else str(file_path)
        
        self.if_print = if_print
        self.if_backup = if_backup

        self._mode = ""

        self._read_file = None
        self._write_file = None
        self._read_position = 0
        self.lines = 0

    @property
    def bak_file_path(self):
        return str(self.file_path) + ".bak"
    
    def exists(self):
        return os.path.exists(self.file_path)

    def len(self):
        file_length = 0
        if not self.exists():
            return file_length
        if self.lines == 0:
            with open(self.file_path, 'r', encoding='utf-8') as file:
                while file.readline():
                    file_length+=1
            self.lines = file_length
        return self.lines

    def close(self,mode = "rw"):
        # 关闭文件资源
        if "r" in mode:
            if self._write_file:
                self._write_file.close()
                self._write_file = None
        if "w" in mode:
            if self._read_file:
                self._read_file.close()
                self._read_file = None
            self.lines = 0
        

    def reset(self, file_path:Union[str , pathlib.Path]):
        self.close()
        self.file_path = file_path if not isinstance(file_path,pathlib.Path) else str(file_path)


    def load_line(self,fast:bool=False):
        if not fast:
            if not self.exists():
                rich.print(f"[yellow]{self.file_path}文件不存在,返回{None}")
                return None
            if self._mode != "r":
                self.close("r")
                
        if not self._read_file:
            self._read_file = open(self.file_path, 'r', encoding='utf-8')
            
        if not fast:
            self._read_file.seek(self._read_position)
            self._mode = "r"
       
        try:
            line = self._read_file.readline()
            self._read_position = self._read_file.tell()
            if not line:
                self.close()
                return None
            return json.loads(line.strip())
        except json.JSONDecodeError as e:
            self.close()
            rich.print(f"[red]文件{self.file_path}解析出现错误：{e}")
            return None
        except IOError as e:
            self.close()
            rich.print(f"[red]无法打开文件{self.file_path}：{e}")
            return None
    
    def load_lines(self):
        """获取jsonl中的line，直到结尾"""
        lines = []
        while True:
            line = self.load_line()
            if line ==None:
                break
            lines.append(line)
        return lines
        

    def load_restart(self):
        self.close(mode="r")
        self._read_position = 0
         
    def dump_line(self, data,fast:bool=False):
        if not isinstance(data,dict) and not isinstance(data,list):
            raise ValueError("数据类型不对")
        if not fast:
            # 备份
            if self.len() % 50 == 1 and self.if_backup:
                shutil.copy(self.file_path, self.bak_file_path)
            self._mode = "a"
            # 如果模型尚未打开
        if not self._write_file:
            self._write_file = open(self.file_path, 'a', encoding='utf-8')
        try:
            json_line = json.dumps(data,ensure_ascii=False)
            self._write_file.write(json_line + '\n')
            self._write_file.flush()
            self.lines += 1  
            return True
        except Exception as e:
            
            if os.path.exists(self.bak_file_path) and self.if_backup:
                shutil.copy(self.bak_file_path, self.file_path)
                if self.if_print:
                    rich.print(f"[red]文件{self.file_path}写入失败，已从备份恢复原文件: {e}[/red]")
            else:
                if self.if_print:
                    rich.print(f"[red]文件{self.file_path}写入失败，且无备份可用：{e}[/red]") 
            return False

    def dump_lines(self,datas):
        if not isinstance(datas,list):
            raise ValueError("数据类型不对")
        if self.if_backup and os.path.exists(self.file_path):
            shutil.copy(self.file_path, self.bak_file_path)
        self._mode = "a"
        if not self._write_file:
            self._write_file = open(self.file_path, 'a', encoding='utf-8')
        try:
            self.len()
            for data in datas:
                json_line = json.dumps(data,ensure_ascii=False)
                self._write_file.write(json_line + '\n')
                self.lines += 1  
            self._write_file.flush()
            return True
        except Exception as e:
            if os.path.exists(self.bak_file_path) and self.if_backup:
                shutil.copy(self.bak_file_path, self.file_path)
                if self.if_print:
                    rich.print(f"[red]文件{self.file_path}写入失败，已从备份恢复原文件: {e}[/red]")
            else:
                if self.if_print:
                    rich.print(f"[red]文件{self.file_path}写入失败，且无备份可用：{e}[/red]") 
                return False
            
    def dump_restart(self):
        self.close()
        self._mode= "w"
        with open(self.file_path, 'w', encoding='utf-8') as file:
            pass 
          
    def load(self):
        jsonl_file = []
        if self.exists():
            try:
                with open(self.file_path, 'r', encoding='utf-8') as f:
                    for line in f:
                        jsonl_file.append(json.loads(line))
            except IOError as e:
                rich.print(f"[red]无法打开文件：{e}")
            except json.JSONDecodeError as e:
                rich.print(f"[red]解析 JSON 文件时出错：{e}")
        else:
            rich.print(f"[yellow]{self.file_path}文件不存在，正在传入空文件...[/yellow]")
        return jsonl_file

    def dump(self,jsonl_file:list):
        before_exist = self.exists()
        if self.if_backup and before_exist:
            shutil.copy(self.file_path, self.bak_file_path)
        try:
            self.close()
            self._mode = "w"
            with open(self.file_path, 'w', encoding='utf-8') as f:
                for entry in jsonl_file:
                    json_str = json.dumps(entry,ensure_ascii=False)
                    f.write(json_str + '\n') 
                    self.lines += 1
            if before_exist and self.if_print:
                rich.print(f"[yellow]更新{self.file_path}[/yellow]")
            elif self.if_print:
                rich.print(f"[green]创建{self.file_path}[/green]")
            return True
        except Exception as e:
            if os.path.exists(self.bak_file_path) and self.if_backup:
                shutil.copy(self.bak_file_path, self.file_path)
                if self.if_print:
                    rich.print(f"[red]文件{self.file_path}写入失败，已从备份恢复原文件: {e}[/red]")
            else:
                if self.if_print:
                    rich.print(f"[red]文件{self.file_path}写入失败，且无备份可用：{e}[/red]") 
            return False  

def load_npy_file(file_path:Union[str , pathlib.Path]):
    if isinstance(file_path,pathlib.Path):
        file_path = str(file_path)
    npy_array = np.empty((0,))
    if os.path.exists(file_path):
        try:
            npy_array = np.load(file_path)
        except IOError as e:
            rich.print(f"[red]无法打开文件：{e}[/red]")
    else:
         rich.print(f"[yellow]{file_path}文件不存在，传入np.empty((0,))[/yellow]")

    return npy_array

def dump_npy_file(npy_array:np.ndarray, file_path:Union[str , pathlib.Path],if_print = True):
    if isinstance(file_path,pathlib.Path):
        file_path = str(file_path)
    before_exist = os.path.exists(file_path)
    try:
        np.save(file_path,npy_array)
        if before_exist and if_print:
            rich.print(f"[yellow]更新{file_path}[/yellow]")
        elif if_print:
            rich.print(f"[green]创建{file_path}[/green]")
    except IOError as e:
        rich.print(f"[red]文件写入失败：{e}[/red]")

def load_pickle_file(file_path:Union[str , pathlib.Path]):
    if isinstance(file_path,pathlib.Path):
        file_path = str(file_path)
    pkl_file = {}
    if os.path.exists(file_path):
        try:
            with open(file_path, 'rb') as file:
                # 使用pickle.load加载并反序列化数据
                pkl_file = pickle.load(file)
        except IOError as e:
            rich.print(f"[red]无法打开文件：{e}[/red]")
    else:
         rich.print(f"[yellow]{file_path}文件不存在，传入空文件[/yellow]")

    return pkl_file

def dump_pickle_file(pkl_file, file_path:Union[str , pathlib.Path],if_print = True):
    if isinstance(file_path,pathlib.Path):
        file_path = str(file_path)
    before_exist = os.path.exists(file_path)
    try:
        with open(file_path, 'wb') as file:
            # 使用pickle.dump将数据序列化到文件
            pickle.dump(pkl_file, file)
        if before_exist and if_print:
            rich.print(f"[yellow]更新{file_path}[/yellow]")
        elif if_print:
            rich.print(f"[green]创建{file_path}[/green]")
    except IOError as e:
        rich.print(f"[red]文件写入失败：{e}[/red]")

def load_txt_file(file_path:Union[str , pathlib.Path]):
    if isinstance(file_path,pathlib.Path):
        file_path = str(file_path)
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r') as f:
                txt_file = f.read()
        except IOError as e:
            rich.print(f"[red]无法打开文件：{e}[/red]")
    else:
         rich.print(f"[yellow]{file_path}文件不存在，传入空文件[/yellow]")

    return txt_file

def dump_txt_file(file,file_path:Union[str , pathlib.Path],if_print = True):
    if isinstance(file_path,pathlib.Path):
        file_path = str(file_path)
    before_exist = os.path.exists(file_path)
    try:
        with open(file_path, 'w') as f:
            # 使用pickle.dump将数据序列化到文件
            f.write(str(file))
        if before_exist and if_print:
            rich.print(f"[yellow]更新{file_path}[/yellow]")
        elif if_print:
            rich.print(f"[green]创建{file_path}[/green]")
    except IOError as e:
        rich.print(f"[red]文件写入失败：{e}[/red]")

def load_excel_file_to_dict(file_path:Union[str , pathlib.Path],if_print = True):
    """存储成如下格式：
    {
        "sheet_name1":[
            {
                "column1":"",
                "column2":"",
                "column3":"",
    }]}
    """
    import openpyxl
    if isinstance(file_path,str):
        file_path = pathlib.Path(file_path)
    assert file_path.suffix == ".xlsx"
    wb = openpyxl.load_workbook(file_path)
    data = {}
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        sheet_data = []

        for row in rows[1:]:
            row_data = {headers[i]: row[i] for i in range(len(headers))}
            sheet_data.append(row_data)

        data[sheet] = sheet_data
    return data

def dump_excel_file(file:dict, file_path:Union[str , pathlib.Path],if_print = True):
    """转换各种模式为xlsx(excel模式)"""
    import openpyxl
    if isinstance(file_path,str):
        file_path = pathlib.Path(file_path)
    assert file_path.suffix == ".xlsx"
    
    wb = openpyxl.Workbook()
    
    if isinstance(file, dict):
        """
        如果是dict，暂时要求符合如下格式：
        {
            "sheet_name1":[
                {
                    "column1":"",
                    "column2":"",
                    "column3":"",
        }]}
        """

        sheet0 = list(file.values())[0]
        assert isinstance(sheet0, list)
        row0 = sheet0[0]
        assert isinstance(row0,dict)
        item0 = list(row0.values())[0]
        assert isinstance(item0,str)
        # 然后转成DataFrame模式
        wb.remove(wb.active)  # 移除默认创建的空白工作表
        # 遍历 JSON 数据中的每个工作表
        for sheet_name, rows in file.items():
            ws = wb.create_sheet(title=sheet_name)  # 创建新的工作表
            headers = rows[0].keys()  # 假设所有行的键相同，作为表头
            ws.append(list(headers))  # 添加表头
            for row in rows:
                ws.append(list(row.values()))  # 添加数据行
    try:
        wb.save(file_path)
    except IOError as e:
        rich.print(f"[red]文件写入失败：{e}[/red]")
        
    if file_path.exists() and if_print:
        rich.print(f"[yellow]更新{file_path}[/yellow]")
    elif if_print:
        rich.print(f"[green]创建{file_path}[/green]")
